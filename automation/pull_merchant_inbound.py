"""
Merchant Services inbound call analytics — daily, Jan 2025 to present.

Runs in GitHub Actions (creds from secrets). Finds queues whose name
matches QUEUE_FILTER (default: "merchant"), pulls daily inbound voice
aggregates for those queues, resolves ALL agent ids (including inactive/
deleted users = past employees), and writes an Excel workbook + CSVs to
./output/ for artifact upload.

Env:
    GENESYS_CLIENT_ID / GENESYS_CLIENT_SECRET / GENESYS_REGION
    QUEUE_FILTER  (optional, default "merchant"; comma-separated terms OK)

If no queue matches, prints the full queue list and exits so you can
re-run the workflow with the right filter term.
"""
import csv
import json
import os
import sys
from datetime import date, datetime, timedelta

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REGION = os.environ.get("GENESYS_REGION", "usw2.pure.cloud")
CID = os.environ.get("GENESYS_CLIENT_ID", "")
SECRET = os.environ.get("GENESYS_CLIENT_SECRET", "")
QUEUE_FILTER = [t.strip().lower() for t in
                os.environ.get("QUEUE_FILTER", "merchant").split(",") if t.strip()]
START = date(2025, 1, 1)
TZ = "America/Los_Angeles"

if not CID or not SECRET:
    print("Missing GENESYS_CLIENT_ID / GENESYS_CLIENT_SECRET"); sys.exit(1)

API = f"https://api.{REGION}"
r = requests.post(f"https://login.{REGION}/oauth/token",
                  data={"grant_type": "client_credentials"},
                  auth=(CID, SECRET), timeout=30)
r.raise_for_status()
H = {"Authorization": f"Bearer {r.json()['access_token']}",
     "Content-Type": "application/json"}
print("AUTH OK")

# ── Queues: list all, match filter ──────────────────────────────────────────
queues = {}
page = 1
while True:
    resp = requests.get(f"{API}/api/v2/routing/queues",
                        params={"pageSize": 100, "pageNumber": page},
                        headers=H, timeout=30)
    resp.raise_for_status()
    body = resp.json()
    for q in body.get("entities", []):
        queues[q["id"]] = q["name"]
    if page >= (body.get("pageCount") or 1): break
    page += 1

matched = {qid: name for qid, name in queues.items()
           if any(t in name.lower() for t in QUEUE_FILTER)}
print(f"\nAll queues ({len(queues)}):")
for name in sorted(queues.values()):
    print(f"  {'-> ' if name in matched.values() else '   '}{name}")
if not matched:
    print(f"\nNO queue matched filter {QUEUE_FILTER}. Re-run the workflow "
          f"with a 'queue_filter' input matching one of the names above.")
    sys.exit(0)
print(f"\nMatched {len(matched)} queue(s): {list(matched.values())}")

# ── Users: full directory incl. inactive/deleted (past employees) ───────────
users = {}
page = 1
while True:
    resp = requests.get(f"{API}/api/v2/users",
                        params={"pageSize": 100, "pageNumber": page,
                                "state": "any"},
                        headers=H, timeout=30)
    resp.raise_for_status()
    body = resp.json()
    for u in body.get("entities", []):
        users[u["id"]] = {"name": u.get("name", u["id"]),
                          "email": u.get("email", ""),
                          "state": u.get("state", "")}
    if page >= (body.get("pageCount") or 1): break
    page += 1
print(f"User directory loaded: {len(users)} users (incl. inactive/deleted)")

# ── Aggregate queries: monthly chunks, daily granularity ────────────────────
def month_chunks(start, end):
    cur = date(start.year, start.month, 1)
    while cur <= end:
        nxt = date(cur.year + (cur.month == 12), (cur.month % 12) + 1, 1)
        yield cur, min(nxt, end + timedelta(days=1))
        cur = nxt

def agg_query(interval, group_by, metrics, queue_ids):
    body = {
        "interval": interval, "granularity": "P1D", "timeZone": TZ,
        "groupBy": group_by, "metrics": metrics,
        "filter": {"type": "and", "clauses": [
            {"type": "or", "predicates": [
                {"dimension": "queueId", "value": q} for q in queue_ids]},
            {"type": "and", "predicates": [
                {"dimension": "mediaType", "value": "voice"},
                {"dimension": "direction", "value": "inbound"}]},
        ]},
    }
    resp = requests.post(f"{API}/api/v2/analytics/conversations/aggregates/query",
                         headers=H, data=json.dumps(body), timeout=60)
    resp.raise_for_status()
    return resp.json().get("results", [])

by_queue = {}   # (date, queue) -> {offered, answered, abandon_n, talk_s}
by_agent = {}   # (date, agent) -> {answered, talk_s}
today = date.today()
qids = list(matched)

for cs, ce in month_chunks(START, today):
    interval = f"{cs.isoformat()}T00:00:00/{ce.isoformat()}T00:00:00"
    # Queue-level: offered / answered / abandoned
    for res in agg_query(interval, ["queueId"],
                         ["nOffered", "nConnected", "tTalkComplete", "tAbandon"],
                         qids):
        qname = matched.get(res.get("group", {}).get("queueId"), "?")
        for bucket in res.get("data", []):
            day = bucket["interval"][:10]
            row = by_queue.setdefault((day, qname),
                                      {"offered": 0, "answered": 0,
                                       "abandoned": 0, "talk_s": 0.0})
            for m in bucket.get("metrics", []):
                st = m.get("stats", {})
                if m["metric"] == "nOffered":
                    row["offered"] += int(st.get("count", 0))
                elif m["metric"] == "nConnected":
                    row["answered"] += int(st.get("count", 0))
                elif m["metric"] == "tAbandon":
                    row["abandoned"] += int(st.get("count", 0))
                elif m["metric"] == "tTalkComplete":
                    row["talk_s"] += float(st.get("sum", 0)) / 1000.0
    # Agent-level: answered + talk time
    for res in agg_query(interval, ["queueId", "userId"],
                         ["nConnected", "tTalkComplete"], qids):
        uid = res.get("group", {}).get("userId")
        if not uid: continue
        u = users.get(uid, {"name": uid, "email": "", "state": "?"})
        for bucket in res.get("data", []):
            day = bucket["interval"][:10]
            row = by_agent.setdefault((day, u["name"]),
                                      {"answered": 0, "talk_s": 0.0,
                                       "email": u["email"], "state": u["state"]})
            for m in bucket.get("metrics", []):
                st = m.get("stats", {})
                if m["metric"] == "nConnected":
                    row["answered"] += int(st.get("count", 0))
                elif m["metric"] == "tTalkComplete":
                    row["talk_s"] += float(st.get("sum", 0)) / 1000.0
    print(f"  chunk {cs} .. {ce - timedelta(days=1)}: "
          f"queue-days={len(by_queue)} agent-days={len(by_agent)}")

# ── Output ──────────────────────────────────────────────────────────────────
os.makedirs("output", exist_ok=True)
stamp = f"2025-01_to_{today.isoformat()}"
xlsx_path = f"output/merchant_services_inbound_{stamp}.xlsx"

wb = Workbook()
HDRF = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HFILL = PatternFill("solid", start_color="1F5577")
BODYF = Font(name="Arial", size=10)

def sheet_with(ws, headers, rows, widths):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HDRF; c.fill = HFILL
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    for r_i, row in enumerate(rows, 2):
        for c_i, v in enumerate(row, 1):
            c = ws.cell(row=r_i, column=c_i, value=v)
            c.font = BODYF
            if isinstance(v, float):
                c.number_format = "#,##0.00"
    ws.freeze_panes = "A2"

ws = wb.active; ws.title = "Daily by Agent"
sheet_with(ws,
    ["Date", "Agent", "Email", "User State", "Inbound Answered",
     "Inbound Talk Time (hrs)"],
    [[d, a, v["email"], v["state"], v["answered"], round(v["talk_s"]/3600, 2)]
     for (d, a), v in sorted(by_agent.items())],
    [12, 24, 32, 11, 16, 20])

ws = wb.create_sheet("Daily by Queue")
sheet_with(ws,
    ["Date", "Queue", "Inbound Offered", "Inbound Answered",
     "Inbound Abandoned", "Talk Time (hrs)"],
    [[d, q, v["offered"], v["answered"], v["abandoned"],
      round(v["talk_s"]/3600, 2)]
     for (d, q), v in sorted(by_queue.items())],
    [12, 28, 15, 16, 17, 15])

# Monthly rollup per agent (computed — static export, no formulas needed)
monthly = {}
for (d, a), v in by_agent.items():
    key = (d[:7], a)
    m = monthly.setdefault(key, {"answered": 0, "talk_s": 0.0})
    m["answered"] += v["answered"]; m["talk_s"] += v["talk_s"]
ws = wb.create_sheet("Monthly by Agent")
sheet_with(ws,
    ["Month", "Agent", "Inbound Answered", "Inbound Talk Time (hrs)"],
    [[mo, a, v["answered"], round(v["talk_s"]/3600, 2)]
     for (mo, a), v in sorted(monthly.items())],
    [10, 24, 16, 20])

ws = wb.create_sheet("Info")
info = [
    ["Report", "Merchant Services inbound call analytics, daily"],
    ["Window", f"{START.isoformat()} through {today.isoformat()}"],
    ["Timezone", TZ],
    ["Queues included", "; ".join(sorted(matched.values()))],
    ["Queue filter term(s)", ", ".join(QUEUE_FILTER)],
    ["Generated", datetime.now().isoformat(timespec="seconds")],
    ["Notes", "Inbound voice only, filtered to the queues above. "
              "Agent rows include past employees (inactive/deleted users). "
              "Queue-level Offered/Abandoned cannot be attributed to agents. "
              "History is queried against current queue IDs — if the team "
              "used a different/renamed queue earlier in the window, those "
              "months will be empty; re-run with a broader queue_filter. "
              "Direct-to-agent (DID) inbound calls that bypass the queue "
              "are not included."],
]
for r_i, (k, v) in enumerate(info, 1):
    a = ws.cell(row=r_i, column=1, value=k); a.font = Font(name="Arial", bold=True, size=10)
    b = ws.cell(row=r_i, column=2, value=v); b.font = BODYF
    b.alignment = Alignment(wrap_text=True, vertical="top")
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 110

wb.save(xlsx_path)

for name, data, headers in [
    (f"output/daily_by_agent_{stamp}.csv", by_agent,
     ["Date", "Agent", "Email", "User State", "Inbound Answered", "Talk Hrs"]),
    (f"output/daily_by_queue_{stamp}.csv", by_queue,
     ["Date", "Queue", "Offered", "Answered", "Abandoned", "Talk Hrs"]),
]:
    with open(name, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f); w.writerow(headers)
        for (d, k2), v in sorted(data.items()):
            if "offered" in v:
                w.writerow([d, k2, v["offered"], v["answered"], v["abandoned"],
                            round(v["talk_s"]/3600, 2)])
            else:
                w.writerow([d, k2, v["email"], v["state"], v["answered"],
                            round(v["talk_s"]/3600, 2)])

print(f"\nWROTE {xlsx_path}")
print(f"Agent-day rows: {len(by_agent)} | Queue-day rows: {len(by_queue)}")
